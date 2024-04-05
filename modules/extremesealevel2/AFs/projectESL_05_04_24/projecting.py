'''
@author: Tim Hermans
t(dot)h(dot)j(dot)hermans@uu(dot)nl
'''
import numpy as np
import geopandas as geopd
from shapely.geometry import Point
from utils import kmdist
from tqdm import tqdm
import os
from openpyxl import load_workbook

def compute_AFs(f,z_hist,slr,refFreq):
    '''computes amplification factor based on historical return curve(s) "z_hist" = F("f"), sea-level projections "slr" at slr.years and reference frequency "refFreq"'''    
    i_ref = np.argmin(np.abs(f-refFreq)) #find index of f closest to refFreq
    
    if (z_hist.ndim == 1) & (len(z_hist)==len(f)): #expand z_hist if only dimension is f (no samples)
        if np.isscalar(slr)==False:
            z_hist = np.repeat(z_hist[:,None],len(slr.samples),axis=1)
    
    i_z_min = np.nanargmin(z_hist,axis=0)[0] #find index of lowest defined z (highest defined frequency); this determines upper limit of AF. Same for all samples because no uncertainty in lower bound is considered.
    
    #z_fut = z_hist+slr
    z_hist_ = np.repeat(z_hist[...,None],len(slr.years),axis=-1) #expand z_hist matrix 
    z_fut = z_hist_ + slr.expand_dims({'f':f}) #compute future return curves by adding SLR for each frequency
    
    refZ_hist = z_hist_[i_ref] #historical reference height samples corresponding to refFreq
    
    if np.isnan(refZ_hist).all(): # when refFreq is higher or lower than where z_hist is defined (for all samples or for none because lower bound has no uncertainty), return nans   
        return [np.nan],np.nan,z_fut 
        
    iRefZ_hist_in_z_fut = np.nanargmin(np.abs(z_fut - refZ_hist),axis=0) #find future frequency corresponding to those heights
    
    AF = f[iRefZ_hist_in_z_fut]/refFreq #divide future frequency correspondong to refZ_hist by historical reference frequency (will evaluate to nan if refFreq=nan)
    
    max_AF = f[i_z_min]/f[i_ref] #keep track of what AF could maximally be given the return curves
    
    return AF,max_AF,z_fut


def compute_AF_timing(f,z_hist,slr,refFreq,AF):
    if AF<=1:
        print("Warning: requested AF<=1; returning projected timing=nan")
        return np.zeros(len(slr.samples)) * np.nan
    
    if np.isnan(refFreq): #refFreq undefined, for instance if location not sufficiently near protection information
        return np.zeros(len(slr.samples)) * np.nan
    
    slr_years=slr.years.values
    
    i_ref = np.argmin(np.abs(f-refFreq)) #find index of f closest to refFreq
    i_ref_AF = np.argmin(np.abs(f-refFreq*AF)) #find index of f closest to AF * refFreq
   
    if (z_hist.ndim == 1) & (len(z_hist)==len(f)): #expand z_hist if only dimension is f (no samples)
        z_hist = np.repeat(z_hist[:,None],len(slr.samples),axis=1)
        
    req_slr = z_hist[i_ref] - z_hist[i_ref_AF] #sea-level rise required to go from refFreq to AF*refFreq (#if one of these is undefined, that's the case for all samples?)
    
    if np.isnan(req_slr).all(): #reference return height and/or target return height not supported by z_hist
        return np.zeros(len(slr.samples)) * np.nan
    
    #find the first timestep at which SLR > required SLR:
    slr_minus_required = slr - np.repeat(req_slr[:,np.newaxis],len(slr_years),axis=1) #projected minus required sea-level rise at each timestep
    slr_minus_required = np.where(slr_minus_required<0,999,slr_minus_required) #minimize only consider where SLR > required SLR (so set negative to 999)
    
    imin = np.nanargmin(slr_minus_required,axis=-1) #find earliest timestep where SLR > required SLR

    #interpolate between earliest timestep and timestep before that (but use earliest timestep if it is the first timestep of the projections):
    slr_years_ = np.hstack((slr_years[0],slr_years)) #copy SLR info and preprend first timestep
    slr_ = np.hstack((np.tile(slr[:, [0]], 1),slr))
    
    #calculate dT/dZ slope
    timing = slr_years[imin]
    timing_tmin1 = slr_years_[imin]
    
    dTime = timing - timing_tmin1
    dZ = slr_[np.arange(len(slr.samples)),imin+1] - slr_[np.arange(len(slr.samples)),imin]
    
    #interpolate by subtracting dT/dZ * (SLR-required_SLR) from earliest timestep with SLR>required SLR
    timing = timing - dTime/(dZ+1e-8) * slr_minus_required[np.arange(len(slr.samples)),imin]
    
    timing = np.where(slr.isel(years=-1)<req_slr,9999,timing) #where not before 2150, set to 9999 (so that it counts as > last timestep in computing quantiles)
    
    return timing


def find_flopros_protection_levels(qlons,qlats,flopros_dir,maxdist):
    '''find flood protection level of polygons nearest (within "maxdist") to queried sites (qlons,qlats)'''
    polygons = geopd.read_file(os.path.join(flopros_dir,'Results_adaptation_objectives/Countries_States_simplified.shp')) #shape file with region polygons
    wb = load_workbook(filename = os.path.join(flopros_dir,'FLOPROS_geogunit_107.xlsx')) #protection standards for each region
    ws = wb.active
    flopros = np.array([cell.value for cell in ws['D'][1::]],dtype=float)
    
    centroids = polygons.centroid #determine polygon centroids for quick filtering
    latc = [k.coords.xy[-1][0] for k in centroids]
    lonc = [k.coords.xy[0][0] for k in centroids]
    
    nearest_segments = []
    
    polygons = polygons.set_crs('epsg:4326')
    
    for qlat,qlon in tqdm(zip(qlats,qlons)): #loop over query coordinates
        p = Point(qlon,qlat) #put coords into point geometry
        
        #do a first filtering based on angular distance to segment/polygon center points (more efficient than shapely distance)
        kmdists = kmdist(qlat,qlon,latc,lonc)
        nearby = np.where(kmdists<50)[0]
        if len(nearby)==0:
            nearest_segments.append(np.nan)
            #add a warning?
            continue
        
        i_near = polygons.iloc[nearby].index[np.argsort( [p.distance(k) for k in polygons.iloc[nearby].geometry])[0:3]] #find the nearest 3 segments based on euclidean distances (default for shapely)
        
        #for these segments, refine the distance computation using a coordinate-specific appropriate reference system
        p_gdf = geopd.GeoDataFrame(geometry=[p], crs='epsg:4326') #put query point into cartesian geodataframe
        if qlat<0: #degermine appropriate epsg system
            e = 32701 + np.floor_divide((qlon+180),6)
            if qlat<-80:
                e = 32761
        else:
            e = 32601 + np.floor_divide((qlon+180),6)
            if qlat>80:
                e = 32661
            
        p_gdf.to_crs(epsg=e, inplace=True) #convert point to epsg system
        polygons_ = polygons.iloc[i_near].to_crs(epsg=e) #convert coordinates of 3 nearest diva segments to epsg system
        
        km_dists = np.array([p_gdf.distance(k).values[0] for k in polygons_.geometry])/1000 #compute kilometric distances
        
        if np.min(km_dists>maxdist): #if nearest point on nearest segment is more than 15km away
            nearest_segments.append(np.nan)
            #add a warning?
            continue
    
        nearest_segments.append(polygons_.index[np.argmin( km_dists )]) #append index of nearest diva segment
        
    protection_levels = []
    
    for k in nearest_segments:
        if np.isfinite(k):
            plevel = flopros[polygons.FID_Aque.iloc[k]] #in units of return period
            protection_levels.append(1/plevel) #add protection level to list in units of return frequency
        else:
            protection_levels.append(np.nan)
            
    return np.array(protection_levels)


def find_diva_protection_levels(qlons,qlats,diva_fn,maxdist):
    '''find flood protection level of polygons nearest (within "maxdist") to queried sites (qlons,qlats)'''
    diva = geopd.read_file(diva_fn) #open diva geo file
    
    nearest_segments = []
    
    for qlat,qlon in tqdm(zip(qlats,qlons)): #loop over query coordinates
        p = Point(qlon,qlat) #put coords into point geometry
        
        #do a first filtering based on angular distance to segment/polygon center points (more efficient than shapely distance)
        kmdists = kmdist(qlat,qlon,diva.lati.values,diva.longi.values)
        nearby = np.where(kmdists<50)[0]
        if len(nearby)==0:
            nearest_segments.append(np.nan)
            #add a warning?
            continue
        
        i_near = diva.iloc[nearby].index[np.argsort( [p.distance(k) for k in diva.iloc[nearby].geometry])[0:5]] #find the nearest 5 segments based on euclidean distances (default for shapely)
        
        #for these segments, refine the distance computation using a coordinate-specific appropriate reference system
        p_gdf = geopd.GeoDataFrame(geometry=[p], crs='epsg:4326') #put query point into cartesian geodataframe
        if qlat<0: #degermine appropriate epsg system
            e = 32701 + np.floor_divide((qlon+180),6)
            if qlat<-80:
                e = 32761
        else:
            e = 32601 + np.floor_divide((qlon+180),6)
            if qlat>80:
                e = 32661
            
        p_gdf.to_crs(epsg=e, inplace=True) #convert point to epsg system
        diva_ = diva.iloc[i_near].to_crs(epsg=e) #convert coordinates of 5 nearest diva segments to epsg system
        
        km_dists = np.array([p_gdf.distance(k).values[0] for k in diva_.geometry])/1000 #compute kilometric distances
        
        if np.min(km_dists>maxdist): #if nearest point on nearest segment is more than 15km away
            nearest_segments.append(np.nan)
            #add a warning?
            continue
    
        nearest_segments.append(diva_.index[np.argmin( km_dists )]) #append index of nearest diva segment
    
    protection_levels = []

    for k in nearest_segments:
        if np.isfinite(k):
            plevel = diva.protection_level_modelled.values[k] #in units of return period
            
            if plevel==0: #if no protection, assign protection of 1/2y (following Scussolini et al., 2016)
                plevel = 2
            protection_levels.append(1/plevel) #add protection level to list in units of return frequency
        else:
            protection_levels.append(np.nan)
    return np.array(protection_levels)