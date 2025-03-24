'''
@author: Tim Hermans
t(dot)h(dot)j(dot)hermans@uu(dot)nl
'''

import numpy as np
import xarray as xr
import pandas as pd
import yaml
import os
from utils import mindist, kmdist
from tqdm import tqdm

def load_config(path_to_config):
    '''load projectESL configuration from "cfgpath" '''
    with open(path_to_config, 'r') as f: #open configuration file
        return yaml.safe_load(f)

def open_input_locations(path_to_input,num_samps):
    '''
    Open input input_locations input file.
    
    Input must be a netcdf or zarr file containing lon/lat coords with 'locations' dimension.
    
    Providing sea-level projections as variable 'sea_level_change' with dimensions 'locations', 'years','samples' is optional depending on requested output.
    
    Format following FACTS output.
    '''
    fn = path_to_input
    if fn.endswith('.zarr'):
        input_locations = xr.open_dataset(fn,engine='zarr',chunks='auto')
    elif fn.endswith('.nc'):
        input_locations = xr.open_dataset(fn,chunks='auto')
    else:
        raise Exception('Input_locations input file format not recognized.')

    if 'locations' not in input_locations.dims:
        input_locations = input_locations.expand_dims('locations')
    input_locations['locations'] = input_locations.locations.astype('str') #to use them as keys for dictionary
    
    try:
        input_locations = input_locations.rename({'latitude':'lat','longitude':'lon'})
    except:
        pass
    
    if ('lon' not in input_locations) or ('lat' not in input_locations):
        raise Exception('Input_locations input must contain lon/lat coordinates.')
    
    #load coordinates into memory if not already
    input_locations['lon'] = (['locations'],input_locations['lon'].values)
    input_locations['lat'] = (['locations'],input_locations['lat'].values)
    
    if 'sea_level_change' in input_locations: #determine if file also contains sea-level projections:
        if 'samples' in input_locations.sea_level_change.dims: #randomly select "num_mc" samples
            if num_samps>len(input_locations.samples):
                raise Exception('Insufficient SLR samples for desired number of Monte Carlo samples.')
            else:
                idx = np.sort(np.random.choice(len(input_locations.samples), size=num_samps, replace=False))
            input_locations = input_locations.isel(samples = idx)
        else:
            raise Exception('SLR projections do not contain samples required for uncertainty propagation.')
        
        if input_locations.sea_level_change.units!='m':
            if input_locations.sea_level_change.units=='mm':
                input_locations['sea_level_change'] = input_locations['sea_level_change']/1000
                input_locations.sea_level_change.attrs['units'] = 'm'
            elif input_locations.sea_level_change.units=='cm':
                input_locations['sea_level_change'] = input_locations['sea_level_change']/100
                input_locations.sea_level_change.attrs['units'] = 'm'
            else:
                raise Exception('Unit of sea-level projections not recognized.')
    else:
        print('Warning: no sea-level projections provided for input locations, will not be able to compute AFs.')
    return input_locations

    
def esl_statistics_dict_to_ds(input_locations,esl_statistics):
    '''stores ESL parameters derived from tide gauge records held in "esl_statistics" dictionary at input_locations as .nc file'''
    ds = xr.Dataset(
    data_vars=dict(
        loc=(['locations'], [v['loc'].values[0] for k,v in esl_statistics.items()]),
        scale=(['locations'], [v['scale'].values[0] for k,v in esl_statistics.items()]),
        shape=(['locations'], [v['shape'].values[0] for k,v in esl_statistics.items()]),
        cov=(['locations','i','j'], [v['cov'].values[0] for k,v in esl_statistics.items()]),
        avg_extr_pyear=(['locations'], [v['avg_extr_pyear'].values[0] for k,v in esl_statistics.items()]),
        
    ),
    coords=dict(
        lon=(['locations'], input_locations.sel(locations=list(esl_statistics.keys())).lon.values),
        lat=(['locations'], input_locations.sel(locations=list(esl_statistics.keys())).lat.values),
        locations=input_locations.sel(locations=list(esl_statistics.keys())).locations.values,
        matched_lon=(['locations'], [v['matched_lon'].values[0] for k,v in esl_statistics.items()]),
        matched_lat=(['locations'], [v['matched_lat'].values[0] for k,v in esl_statistics.items()]),
        matched_id=(['locations'], [v['matched_id'].values[0] for k,v in esl_statistics.items()]),
    ),
    )
   
    try:
        ds['mhhw'] = (['locations'], [v['mhhw'].values[0] for k,v in esl_statistics.items()]) #mhhw is optional in esl_statistics
    except:
        pass
    
    try:
        ds['vdatum'] = (['locations'],[v['vdatum'].values[0] for k,v in esl_statistics.items()])
    except:
        pass
    return ds

def save_ds_to_netcdf(output_path,ds,fn):
    if os.path.exists(output_path) == False:
        os.mkdir(output_path)
    return ds.to_netcdf(os.path.join(output_path,fn),mode='w')

def lazy_output_to_ds(output,f,out_qnts,esl_statistics,target_years=None,target_AFs=None,target_freqs=None): #may need to find a better name?
    
    output_ds = xr.Dataset(data_vars=dict(),
                           coords=dict(f=(['f'],f),
                                       qnt = (out_qnts),
                                       locations = esl_statistics.locations,
                                       lon = esl_statistics.lon,
                                       lat = esl_statistics.lat))
    
    output_ds['z_hist'] = (['locations','qnt','f'],np.stack([k[0] for k in output]))
    
    if len(target_years)>0:
        output_ds['z_fut'] = (['locations','qnt','f','target_year'],np.stack([k[1] for k in output]))
        output_ds['AF'] = (['locations','qnt','target_year'],np.stack([k[2] for k in output]))
        output_ds['maxAF'] = (['locations'],np.stack([k[3] for k in output]))
        output_ds = output_ds.assign_coords({'target_year':target_years})
        
    if len(target_AFs)>0:
        output_ds['AF_timing'] = (['locations','qnt','target_AF'],np.stack([k[4] for k in output]))
        output_ds = output_ds.assign_coords({'target_AF':target_AFs})

    if len(target_freqs)>0:
        output_ds['f_timing'] = (['locations','qnt','target_f'],np.stack([k[5] for k in output]))
        output_ds = output_ds.assign_coords({'target_f':target_freqs})
    
    return output_ds


def find_flopros_protection_levels(qlons,qlats,flopros_dir,maxdist):
    '''find flood protection level of polygons nearest (within "maxdist") to queried sites (qlons,qlats)'''
    
    from openpyxl import load_workbook
    import geopandas as geopd
    from shapely.geometry import Point
    
    polygons = geopd.read_file(os.path.join(flopros_dir,'Results_adaptation_objectives/Countries_States_simplified.shp')) #shape file with region polygons
    wb = load_workbook(filename = os.path.join(flopros_dir,'FLOPROS_geogunit_107.xlsx')) #protection standards for each region
    ws = wb.active
    flopros = np.array([cell.value for cell in ws['D'][1::]],dtype=float)
    
    centroids = polygons.centroid #determine polygon centroids for quick filtering
    latc = [k.coords.xy[-1][0] for k in centroids]
    lonc = [k.coords.xy[0][0] for k in centroids]
    
    nearest_segments = []
    
    polygons = polygons.set_crs('epsg:4326')
    
    for qlat,qlon in tqdm(zip(qlats,qlons)): #loop over query coordinates
        p = Point(qlon,qlat) #put coords into point geometry
        
        #do a first filtering based on angular distance to segment/polygon center points (more efficient than shapely distance)
        kmdists = kmdist(qlat,qlon,latc,lonc)
        nearby = np.where(kmdists<5000)[0]
        if len(nearby)==0:
            nearest_segments.append(np.nan)
            #add a warning?
            continue
        
        i_near = polygons.iloc[nearby].index[np.argsort( [p.distance(k) for k in polygons.iloc[nearby].geometry])[0:3]] #find the nearest 3 segments based on euclidean distances (default for shapely)
        
        #for these segments, refine the distance computation using a coordinate-specific appropriate reference system
        p_gdf = geopd.GeoDataFrame(geometry=[p], crs='epsg:4326') #put query point into cartesian geodataframe
        if qlat<0: #determine appropriate epsg system
            e = 32701 + np.floor_divide((qlon+180),6)
            if qlat<-80:
                e = 32761
        else:
            e = 32601 + np.floor_divide((qlon+180),6)
            if qlat>80:
                e = 32661
            
        p_gdf.to_crs(epsg=e, inplace=True) #convert point to epsg system
        polygons_ = polygons.iloc[i_near].to_crs(epsg=e) #convert coordinates of 3 nearest diva segments to epsg system
        
        km_dists = np.array([p_gdf.distance(k).values[0] for k in polygons_.geometry])/1000 #compute kilometric distances
        
        if np.min(km_dists>maxdist): #if nearest point on nearest segment is more than max dist km away
            nearest_segments.append(np.nan)
            #add a warning?
            continue
    
        nearest_segments.append(polygons_.index[np.argmin( km_dists )]) #append index of nearest diva segment
        
    protection_levels = []
    
    for k in nearest_segments:
        if np.isfinite(k):
            plevel = flopros[polygons.FID_Aque.iloc[k]] #in units of return period
            protection_levels.append(1/plevel) #add protection level to list in units of return frequency
        else:
            protection_levels.append(np.nan)
            
    return np.array(protection_levels)


def find_diva_protection_levels(qlons,qlats,diva_fn,maxdist):
    '''find flood protection level of polygons nearest (within "maxdist") to queried sites (qlons,qlats)'''
    import geopandas as geopd
    from shapely.geometry import Point
    
    diva = geopd.read_file(diva_fn) #open diva geo file
    
    nearest_segments = []
    
    for qlat,qlon in tqdm(zip(qlats,qlons)): #loop over query coordinates
        p = Point(qlon,qlat) #put coords into point geometry
        
        #do a first filtering based on angular distance to segment/polygon center points (more efficient than shapely distance)
        kmdists = kmdist(qlat,qlon,diva.lati.values,diva.longi.values)
        nearby = np.where(kmdists<500)[0]
        if len(nearby)==0:
            nearest_segments.append(np.nan)
            #add a warning?
            continue
        
        i_near = diva.iloc[nearby].index[np.argsort( [p.distance(k) for k in diva.iloc[nearby].geometry])[0:5]] #find the nearest 5 segments based on euclidean distances (default for shapely)
        
        #for these segments, refine the distance computation using a coordinate-specific appropriate reference system
        p_gdf = geopd.GeoDataFrame(geometry=[p], crs='epsg:4326') #put query point into cartesian geodataframe
        if qlat<0: #degermine appropriate epsg system
            e = 32701 + np.floor_divide((qlon+180),6)
            if qlat<-80:
                e = 32761
        else:
            e = 32601 + np.floor_divide((qlon+180),6)
            if qlat>80:
                e = 32661
            
        p_gdf.to_crs(epsg=e, inplace=True) #convert point to epsg system
        diva_ = diva.iloc[i_near].to_crs(epsg=e) #convert coordinates of 5 nearest diva segments to epsg system
        
        km_dists = np.array([p_gdf.distance(k).values[0] for k in diva_.geometry])/1000 #compute kilometric distances
        
        if np.min(km_dists>maxdist): #if nearest point on nearest segment is more than 15km away
            nearest_segments.append(np.nan)
            #add a warning?
            continue
    
        nearest_segments.append(diva_.index[np.argmin( km_dists )]) #append index of nearest diva segment
    
    protection_levels = []

    for k in nearest_segments:
        if np.isfinite(k):
            plevel = diva.protection_level_modelled.values[k] #in units of return period
            
            if plevel==0: #if no protection, assign protection of 1/2y (following Scussolini et al., 2016)
                plevel = 2
            protection_levels.append(1/plevel) #add protection level to list in units of return frequency
        else:
            protection_levels.append(np.nan)
    return np.array(protection_levels)

def find_custom_protection_levels(qlons,qlats,custom_fn,maxdist):
    #custom protection must be given as xarray dataset, with coordinates "lon", "lat", "locations", and variable "protection" in 1/yr as funtion of "locations"
    ds = xr.open_dataset(custom_fn)
    
    ds_lons = ds.lon.values
    ds_lats = ds.lat.values
    
    min_idx = [mindist(x,y,ds_lats,ds_lons,maxdist) for x,y in zip(qlats,qlons)]

    protection_levels = np.zeros((len(qlats)))
    
    for i in np.arange(len(protection_levels)): #for each input location
        if min_idx[i] is not None: #if nearby ESL information found
            try:
                j = min_idx[i][0] #if multiple within maxdist, take the first
            except:
                j = min_idx[i]
            protection_levels[i] = ds.protection.isel(locations=j).values
        else:
            protection_levels[i] = np.nan
   
    return protection_levels
    
def get_refFreqs(refFreq_data,input_locations,esl_statistics,path_to_refFreqs=None):
    ''' determine reference frequencies for queried input_locations based on user options in config'''
    
    qlats = input_locations.lat.sel(locations=esl_statistics.locations).values
    qlons = input_locations.lon.sel(locations=esl_statistics.locations).values
    
    if refFreq_data == 'diva': #find contemporary DIVA flood protection levels
        refFreqs = find_diva_protection_levels(qlons,qlats,path_to_refFreqs,10) #set these paths in config
    
    elif refFreq_data == 'flopros': #find contemporary FLOPROS flood protection levels
        refFreqs = find_flopros_protection_levels(qlons,qlats,path_to_refFreqs,10)    
        
    elif refFreq_data == 'custom':
        refFreqs = find_custom_protection_levels(qlons,qlats,path_to_refFreqs,10)
        
    elif np.isscalar(refFreq_data): #use constant reference frequency for every location
            refFreqs = np.tile(refFreq_data,len(esl_statistics.locations))
    else: 
        raise Exception('Rquested reference frequency must be "diva", "flopros","custom" or a constant.')
            
    return refFreqs

def open_gpd_parameters(input_data,data_path,input_locations,n_samples,match_dist_limit):
    if input_data == 'hermans2023':
        gpd_params = xr.open_dataset(data_path) #open GPD parameters
        min_idx = [mindist(x,y,gpd_params.lat.values,gpd_params.lon.values, match_dist_limit) for x,y in zip(input_locations.lat.values, input_locations.lon.values)] #find ESL indices within radius of input locations if any
        
        if len(gpd_params.nboot) > n_samples: #determine which samples to select
            isamps = np.random.randint(0,len(gpd_params.nboot),n_samples)
        elif len(gpd_params.nboot) < n_samples:
            isamps = np.hstack((gpd_params.nboot.values,np.random.randint(0,len(gpd_params.nboot),n_samples-len(gpd_params.nboot))))
        else:
            isamps = gpd_params.nboot.values
        
        #generate lists with indices to keep
        iLocations = [] 
        iEsl = []
        for i in np.arange(len(input_locations.locations)): #for each input location
            if min_idx[i] is not None: #if nearby ESL information found
                if np.isscalar(min_idx[i]) == False: #if multiple, pick the first (we don't have information about series length here)
                    iEsl.append(min_idx[i][0])
                else:
                    iEsl.append(min_idx[i]) #if only one, use that
                iLocations.append(i) #nearby ESL information for this site, so store site
            else: #no nearby ESL information found
                print("Warning: No nearby ESL information found for {0}. Skipping site.".format(input_locations.locations[i].values))
                continue
        
        #put selected information in new dataset
        esl_statistics = gpd_params.isel(site=iEsl).isel(nboot=isamps) #selected sites & samples
        esl_statistics = esl_statistics.rename({'avg_exceed':'avg_extr_pyear','nboot':'samples','location':'loc','site':'locations'}) #rename some variables
        esl_statistics = esl_statistics[['loc','avg_extr_pyear','scale','shape','scale_samples','shape_samples']]
        esl_statistics['locations'] = input_locations.isel(locations=iLocations).locations
        esl_statistics = esl_statistics.assign_coords({'lon':input_locations.isel(locations=iLocations).lon,
                                                       'lat':input_locations.isel(locations=iLocations).lat,
                                                       'matched_lon':('locations',gpd_params.isel(site=iEsl).lon.values),
                                                       'matched_lat':('locations',gpd_params.isel(site=iEsl).lat.values),
                                                       'matched_id':('locations',gpd_params.isel(site=iEsl).site.values)})
        
    elif input_data == 'kirezci2020':
        from esl_analysis import infer_avg_extr_pyear
        '''open GPD parameters from Kirezci et al. (2020) and find parameters nearest to queried input_locations'''
        gpd_params = pd.read_csv(data_path)
    
        min_idx = [mindist(x,y,gpd_params.lat.values,gpd_params.lon.values,match_dist_limit) for x,y in zip(input_locations.lat.values, input_locations.lon.values)]
    
        #generate lists with indices to keep
        iLocations = [] 
        iEsl = []
        for i in np.arange(len(input_locations.locations)): #for each input location
            if min_idx[i] is not None: #if nearby ESL information found
                if np.isscalar(min_idx[i]) == False: #if multiple, pick the first (we don't have information about series length here)
                    iEsl.append(min_idx[i][0])
                else:
                    iEsl.append(min_idx[i]) #if only one, use that
                iLocations.append(i) #nearby ESL information for this site, so store site
            else: #no nearby ESL information found
                print("Warning: No nearby ESL information found for {0}. Skipping site.".format(input_locations.locations[i].values))
                continue
 
              
        esl_statistics = xr.Dataset(data_vars=dict(loc=(['locations'],gpd_params.iloc[iEsl].GPDthresh.values),
                                                   scale=(['locations'],gpd_params.iloc[iEsl].GPDscale.values),
                                                   shape=(['locations'],gpd_params.iloc[iEsl].GPDshape.values),
                                                   avg_extr_pyear=(['locations'],
                                                                   infer_avg_extr_pyear(gpd_params.iloc[iEsl].GPDthresh.values,
                                                                                        gpd_params.iloc[iEsl].GPDscale.values,
                                                                                        gpd_params.iloc[iEsl].GPDshape.values,
                                                                                        gpd_params.iloc[iEsl].TWL.values,
                                                                                        1/100)),))
        esl_statistics['locations'] = input_locations.isel(locations=iLocations).locations
        
        esl_statistics = esl_statistics.assign_coords({'lon':input_locations.isel(locations=iLocations).lon,
                                                       'lat':input_locations.isel(locations=iLocations).lat,
                                                       'matched_lon':('locations',gpd_params.iloc[iEsl].lon.values),
                                                       'matched_lat':('locations',gpd_params.iloc[iEsl].lat.values),
                                                       'matched_id':('locations',iEsl)})    
    elif input_data == 'vousdoukas2018':
        from esl_analysis import infer_avg_extr_pyear
        '''open GPD parameters from Vousdoukas et al. (2018) and find parameters nearest to queried input_locations'''
        gpd_params = xr.open_dataset(data_path).isel(rlyear=0,drop=True)
        
        min_idx = [mindist(x,y,gpd_params.lat.values,gpd_params.lon.values,match_dist_limit) for x,y in zip(input_locations.lat.values, input_locations.lon.values)]
    
        #generate lists with indices to keep
        iLocations = [] 
        iEsl = []
        for i in np.arange(len(input_locations.locations)): #for each input location
            if min_idx[i] is not None: #if nearby ESL information found
                if np.isscalar(min_idx[i]) == False: #if multiple, pick the first (we don't have information about series length here)
                    iEsl.append(min_idx[i][0])
                else:
                    iEsl.append(min_idx[i]) #if only one, use that
                iLocations.append(i) #nearby ESL information for this site, so store site
            else: #no nearby ESL information found
                print("Warning: No nearby ESL information found for {0}. Skipping site.".format(input_locations.locations[i].values))
                continue
        
        gpd_params = gpd_params.rename({'thresholdParamGPD':'loc','scaleParamGPD':'scale','shapeParamGPD':'shape','npt':'locations','pt':'locations'})
        esl_statistics = gpd_params.isel(locations=iEsl)
        matched_ids = esl_statistics.locations.values
        
        esl_statistics['locations'] = input_locations.isel(locations=iLocations).locations.values
        esl_statistics = esl_statistics.assign_coords({'matched_id':('locations',matched_ids),
                                                       'matched_lon':('locations',esl_statistics.lon.values),
                                                       'matched_lat':('locations',esl_statistics.lat.values)})
        esl_statistics['avg_extr_pyear'] = infer_avg_extr_pyear(esl_statistics['loc'],esl_statistics.scale,esl_statistics.shape,esl_statistics.returnlevelGPD.isel(return_period=12),1/100)
        esl_statistics = esl_statistics[['loc','scale','shape','avg_extr_pyear']]
        
        esl_statistics = esl_statistics.assign_coords({'lon':input_locations.isel(locations=iLocations).lon,
                                                       'lat':input_locations.isel(locations=iLocations).lat})
    return esl_statistics


#def if_scalar_to_list(variable):
    
    
def get_coast_rp_return_curves(input_dir,input_locations,f,match_dist_limit):
    '''open return curves from Dulaart et al. (2021) and find curves nearest to queried input_locations'''
    zs = []
    iLocations = []
    iMatched = []
    
    coast_rp_coords = pd.read_pickle(os.path.join(input_dir,'pxyn_coastal_points.xyn'))
    min_idx = [mindist(x,y,coast_rp_coords['lat'].values,coast_rp_coords['lon'].values,match_dist_limit) for x,y in zip(input_locations.lat.values, input_locations.lon.values)]
    #use a slightly larger distance tolerance here because GTSM output is used at locations every 25 km along the global coastline (Dullart et al., 2021).
    
    for i in np.arange(len(input_locations.locations)):
        if min_idx[i] is not None:
            if np.isscalar(min_idx[i]) == False: #if multiple input_locations within radius, pick the first (we don't have information about series length here)
                min_idx[i] = min_idx[i][0]
            this_id = str(int(coast_rp_coords.iloc[min_idx[i]].name)).rjust(5,'0')
            try:
                rc = pd.read_pickle(os.path.join(input_dir,'rp_full_empirical_station_'+this_id+'.pkl'))
            except:
                print("Warning: Could not open {0}. Skipping site.".format('rp_full_empirical_station_'+this_id+'.pkl'))
                continue
            rc =rc['rp'][np.isfinite(rc['rp'])]
            #in some cases coast rp RPs can be non-monotonically increasing, for low heights if tcs defined where etcs not defined; we don't want to use this part
            d = np.where(np.diff(rc)<0)[0]#find where not increasing
            try:
                rc = rc.iloc[d[0]+1::]
            except:
                pass
        
            z_hist = np.flip(rc.index.values)
            f_hist = np.flip(1/rc.values)
        
            z = np.interp(f,f_hist,z_hist,left=np.nan,right=np.nan)
            
            zs.append(z)
            iLocations.append(i)
            iMatched.append(min_idx[i])
        else:
            print("Warning: No nearby ESL information found for {0}. Skipping site.".format(input_locations.locations[i].values))
            continue
        
    esl_statistics = xr.Dataset(data_vars=dict(z_hist=(['locations','f'],np.stack(zs))),
                                coords=dict(locations=input_locations.isel(locations=iLocations).locations,
                                            f=f))
    esl_statistics = esl_statistics.assign_coords({'matched_lon':('locations',coast_rp_coords.iloc[iMatched].lon.values),
                                                   'matched_lat':('locations',coast_rp_coords.iloc[iMatched].lat.values),
                                                   'matched_id':('locations',coast_rp_coords.iloc[iMatched].station.values),
                                                    'lon':input_locations.isel(locations=iLocations).lon,
                                                   'lat':input_locations.isel(locations=iLocations).lat})
    
    return esl_statistics


def open_gtsm_waterlevels(gtsm_dir, input_locations, match_dist_limit):
    gtsm = xr.open_mfdataset(os.path.join(gtsm_dir,'*.nc')) #assumes daily maxima in similar file structure as original CDS download
    gtsm_lats = gtsm.station_y_coordinate.values
    gtsm_lons = gtsm.station_x_coordinate.values
    
    min_idx = [mindist(x,y,gtsm_lats,gtsm_lons,match_dist_limit) for x,y in zip(input_locations.lat.values, input_locations.lon.values)]
    for k in np.arange(len(input_locations.locations)):
        if min_idx[k] is not None:
            min_idx[k] = min_idx[k][0]
        else:
            min_idx[k] = np.nan
            print("Warning: No nearby ESL information found for {0}. Skipping site.".format(input_locations.locations[k].values))
    matched_ids = gtsm.stations.values[np.array(min_idx)[np.isfinite(min_idx)].astype('int')]
    gtsm = gtsm.isel(stations = np.array(min_idx)[np.isfinite(min_idx)].astype('int'))
    gtsm = gtsm.rename({'stations':'locations','station_x_coordinate':'matched_lon','station_y_coordinate':'matched_lat'})
    gtsm = gtsm.assign_coords({'matched_id':('locations',matched_ids),'lon':input_locations.lon[np.isfinite(min_idx)],'lat':input_locations.lat[np.isfinite(min_idx)]})
    gtsm['locations'] =  input_locations.locations[np.isfinite(min_idx)]
        
    gtsm = gtsm.load() #load data into memory
    
    #do some cleaning up (not entirely sure why this is necessary, also seems to be the case in the original dataset)
    gtsm = gtsm.drop_isel(locations = np.where(np.isnan(gtsm.waterlevel).any(dim='time'))[0]) #remove stations containing nans in their timeseries
    gtsm = gtsm.where((gtsm.waterlevel.var(dim='time')>1e-5)).dropna(dim='locations') #remove weird stations with very low variance (erroneous, sea ice, internal seas?)

    return gtsm

